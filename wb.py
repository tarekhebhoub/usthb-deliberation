
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import Workbook

def moyenne(matricule):
        url0='https://ent.usthb.dz/delib/'
        data = {'matricule':matricule,
                'ok':'ok'
                }
        r = requests.post(url = url0, data = data)
        soup = BeautifulSoup(r.content, 'html.parser')
        s = soup.find('div', class_='container')
        moy=soup.find_all('strong')
        if(len(moy)>0):
                moy=moy[len(moy)-1]
                moyenne=str(moy)
                moyenne=moyenne[14:19]
                return moyenne
        else:
                return "00.00"

def list_moyenne(cell_obj,nombre_of_student):
        wb = openpyxl.Workbook()
        sheet = wb.active
        c1=sheet.cell(row=1,column=1)
        c1.value='N°'
        c2 = sheet.cell(row = 1, column = 2)
        c2.value='Matricule'
        c3 = sheet.cell(row= 1 , column = 3)
        c3.value='Moyenne'
        i=2
        while(i<=nombre_of_student):
                c1=sheet.cell(row=i,column=1)
                c1.value=i
                c2 = sheet.cell(row = i, column = 2)
                c2.value=cell_obj.value
                c3 = sheet.cell(row= i , column = 3)
                c3.value=moyenne(cell_obj.value)
                print(str(i)+':'+str(c2.value)+': '+c3.value)
                i+=1
                cell_obj = sheet_obj.cell(row = i, column = 2)
        wb.save("Moyenne_des_etudiants.xlsx")
def max_moyenne(cell_obj,nombre_of_student):
        max=0
        i=1
        while(i<=nombre_of_student):     
                cell_obj = sheet_obj.cell(row = i+1, column = 2)          
                c2 = sheet.cell(row = i+1, column = 1)
                c2.value=cell_obj.value
                moy=moyenne(c2.value)
                x=moy[0]+moy[1]+'.'+moy[3]+moy[4]
                x=float(x)
                if(x>max):
                        max=x
                i+=1
        return max
        
        
# Give the location of the file
path = "list_of_matricule.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet=wb_obj.active
sheet_obj = wb_obj.active
nombre_of_student=int(input("enter the numbre of student: "))
cell_obj = sheet_obj.cell(row = 2, column = 2)
list_moyenne(cell_obj,nombre_of_student)

